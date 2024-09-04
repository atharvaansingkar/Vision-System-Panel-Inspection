# Import necessary libraries
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
import cv2
import numpy as np
import requests
from skimage.exposure import match_histograms
import os
import pandas as pd
from datetime import datetime, timezone, timedelta
from datetime import datetime
import base64
import openpyxl
from openpyxl import Workbook, load_workbook  # Add load_workbook here
from PIL import Image
import logging
import neoapi
import gc
import math
import serial
import time

# Initialize Flask app
app = Flask(__name__)
app.secret_key = '12345678'
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'static')

# Define the credentials
USERNAME = 'admin'
PASSWORD = 'admin'

# Initialize logger
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Function to generate JSON inspection report and handle folder creation
def generate_inspection_report(language, status, image, username):
    try:
        # Create inspection report data
        timestamp_ms = int(datetime.now().timestamp() * 1000)
        current_date = datetime.now().strftime("%d-%m-%Y")
        
        # Define the base directory for saving the images
        base_dir = os.path.join(r'C:\Program Files\Apache Software Foundation\Tomcat 9.0\webapps\sfactory\Panel_Inspection', current_date)
        
        # Create date folder if it doesn't exist
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        
        # Create model name directory inside date folder
        model_dir = os.path.join(base_dir, language)
        if not os.path.exists(model_dir):
            os.makedirs(model_dir)
        
        # Create OK and NG folders inside model name directory
        ok_dir = os.path.join(model_dir, 'OK')
        ng_dir = os.path.join(model_dir, 'NG')
        if not os.path.exists(ok_dir):
            os.makedirs(ok_dir)
        if not os.path.exists(ng_dir):
            os.makedirs(ng_dir)

        # Save the processed image to the appropriate folder based on status
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        result_filename = f"{language}_{timestamp}.jpg"
        status_folder = 'OK' if status == 'OK' else 'NG'
        image_path = os.path.join(model_dir, status_folder, result_filename)
        cv2.imwrite(image_path, image)

        # Also save the result image as result_image.bmp in the static folder
        static_folder_path = r'D:/pisx_test/static'
        static_image_path = os.path.join(static_folder_path, 'result_image.bmp')
        cv2.imwrite(static_image_path, image)

        # Extract the relative path to be sent in the API request
        relative_path = os.path.relpath(image_path, r'C:\Program Files\Apache Software Foundation\Tomcat 9.0\webapps\sfactory')

        # Create the report data
        report_data = {
        "FunctionName": "1007",
        "Name": "MT123",
        "Password": "MT123",
        "UniqueID": 1,
        "DateTime": str(timestamp_ms),
        "UniqueKey": language,  
        "Command": 3,
        "CommandData": {
            "CommandHeader": [
                {"headerkey": "ImagePath", "headervalue": relative_path.replace("\\", "/")},
                {"headerkey": "FinalResult", "headervalue": status},
                {"headerkey": "UserId", "headervalue": username}
            ]
        }
    }

        # Send API data to the desired endpoint
        response = requests.post('http://localhost:8080/sfactory/', json=report_data)
        if response.status_code == 200:
            return response.text  # Return response if needed
        else:
            return None

    except Exception as e:
        logger.error("Error generating inspection report:", exc_info=True)
        return None

# Function to load patch locations from an Excel file
def load_patch_locations(language):
    try:
        base_path = os.path.join('static', 'language_models', language)
        excel_file = os.path.join(base_path, 'location_data.xlsx')
        
        if not os.path.exists(excel_file):
            logger.error(f"Location data file for language {language} does not exist.")
            return [], []

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet1 = wb[wb.sheetnames[0]]
        sheet2 = wb[wb.sheetnames[1]]

        gray_patch_locations_1 = [row[0:4] for row in sheet1.iter_rows(min_row=2, values_only=True)]
        gray_patch_locations_2 = [row[0:4] for row in sheet2.iter_rows(min_row=2, values_only=True)]

        return gray_patch_locations_1, gray_patch_locations_2
    
    except Exception as e:
        logger.error(f"Error loading patch locations for language {language}: {e}", exc_info=True)
        return [], []

# Define the route for capturing image from the camera
@app.route('/capture', methods=['POST'])
def capture_image():
    try:
        camera = neoapi.Cam()
        camera.Connect()
        
        if camera.IsConnected():
            image = camera.GetImage()
            filename = 'captured_image.bmp'
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image.Save(image_path)
            
            return jsonify({'success': True, 'image_path': f'/static/{filename}'})
        else:
            return jsonify({'success': False, 'error': 'Camera not connected'})
    except Exception as e:
        logger.error(f"Error capturing image: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to capture image'})

# File path for the users.xlsx
FILEPATH = 'static/users.xlsx'

# Function to load users from an .xlsx file
def load_users():
    if not os.path.exists(FILEPATH):
        # If the file does not exist, create a new one with headers
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Username', 'Password'])
        wb.save(FILEPATH)
        return {}
    
    wb = openpyxl.load_workbook(FILEPATH)
    sheet = wb.active
    users = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        username, password = row
        users[username] = password
    return users

# Function to save users to an .xlsx file
def save_users(users):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['Username', 'Password'])
    for username, password in users.items():
        sheet.append([username, password])
    wb.save(FILEPATH)

# Load users into a global variable
users = load_users()

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username in users and users[username] == password:
            session['username'] = username
            return redirect(url_for('inspection'))
        else:
            error_message = 'Invalid credentials. Please try again.'
            return render_template('index.html', error_message=error_message)
    return render_template('index.html')


@app.route('/manageusers', methods=['GET'])
def manage_users():
    return render_template('manageusers.html')

@app.route('/add_user', methods=['POST'])
def add_user():
    global users
    new_username = request.form.get('new_username')
    new_password = request.form.get('new_password')
    
    if new_username in users:
        flash('Username already exists.', 'error')
    else:
        users[new_username] = new_password
        save_users(users)
        flash('User added successfully!', 'success')
    return redirect(url_for('manage_users'))

@app.route('/delete_user', methods=['POST'])
def delete_user():
    global users
    delete_username = request.form.get('delete_username')
    
    if delete_username not in users:
        flash('Username not found.', 'error')
    else:
        del users[delete_username]
        save_users(users)
        flash('User deleted successfully!', 'success')
    return redirect(url_for('manage_users'))

@app.route('/inspection')
def inspection():
    if 'username' in session:
        username = session['username']
        return render_template('inspection.html', username=username)
    else:
        return redirect(url_for('index'))
    
@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('index'))   


# Define the route for the Master login page
@app.route('/Masterlogin', methods=['GET', 'POST'])
def Masterlogin():
    error_message = None  # Initialize error message
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == USERNAME and password == PASSWORD:
            # Redirect to another page upon successful login
            return redirect(url_for('master'))
        else:
            error_message = 'Incorrect username or password. Please try again.'

    return render_template('Masterlogin.html', error_message=error_message)

# Define the route for master page (you can replace this with the desired page)
@app.route('/master')
def master():
    return render_template('master.html')

# Define the route for processing images
@app.route('/process', methods=['POST'])
def process_image():
    try:
        # Initialize serial communication with Arduino
        arduino = serial.Serial(port='COM8', baudrate=9600, timeout=1)
        time.sleep(2)  # Give some time for the connection to establish

        # Get the logged-in username from the session
        username = session.get('username', 'Unknown')

        # Get selected language from the request
        language = request.form['language']
        
        # Load the uploaded image
        taken_image = cv2.imread("static/captured_image.bmp", cv2.IMREAD_COLOR)
        uploaded_image = cv2.cvtColor(taken_image, cv2.COLOR_BGR2GRAY)

        # Construct the dynamic path for the reference image
        reference_image_path = os.path.join('static', 'language_models', language, 'reference_image.bmp')

        # Check if the reference image exists
        if not os.path.isfile(reference_image_path):
            return jsonify({"status": "error", "message": "Reference image not found."}), 404
        
        # Load the reference image
        mastered_image = cv2.imread(reference_image_path, cv2.IMREAD_COLOR)
        reference_image = cv2.cvtColor(mastered_image, cv2.COLOR_BGR2GRAY)

        # Read the image to be registered
        image_to_register = uploaded_image

        # Initialize SIFT detector
        sift = cv2.SIFT_create()

        # Detect keypoints and compute descriptors for both images
        keypoints_ref, descriptors_ref = sift.detectAndCompute(reference_image, None)
        keypoints_to_register, descriptors_to_register = sift.detectAndCompute(image_to_register, None)

        # Use FLANN based matcher for finding matches
        flann = cv2.FlannBasedMatcher()
        matches = flann.knnMatch(descriptors_ref, descriptors_to_register, k=2)

        # Apply ratio test to filter good matches
        good_matches = []
        for m, n in matches:
            if m.distance < 0.75 * n.distance:
                good_matches.append(m)

        # Check if there are enough good matches
        if len(good_matches) < 4:
            raise ValueError("Not enough good matches found for reliable homography estimation.")

        # Extract matched keypoints
        matched_keypoints_ref = np.float32([keypoints_ref[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
        matched_keypoints_to_register = np.float32([keypoints_to_register[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)

        # Find homography
        H, _ = cv2.findHomography(matched_keypoints_to_register, matched_keypoints_ref, cv2.RANSAC, 5.0)

        # Warp the image to be registered to align with the reference image
        registered_image = cv2.warpPerspective(image_to_register, H, (reference_image.shape[1], reference_image.shape[0]))

        # Perform histogram matching for grayscale images
        matched = match_histograms(registered_image, reference_image)

        # Template Matching
        def extract_patch(image, x, y, width, height):
            """Function to extract a patch from an image based on the given location and size."""
            patch = image[y:y+height, x:x+width]
            return patch

        # Load patch locations based on the selected language
        gray_patch_locations_1, gray_patch_locations_2 = load_patch_locations(language)

        # Initialize lists to store patch matching results
        gray_patch_matches_1 = []
        gray_patch_matches_2 = []

        # Process patches from gray_patch_locations_1
        for i, (x, y, width, height) in enumerate(gray_patch_locations_1):
            # Extract the patch from the original image
            master_patch = extract_patch(reference_image, x, y, width, height)

            # Extract the patch from the new input image
            test_patch = extract_patch(matched, x, y, width, height)

            # Ensure the patches are of the same type and convert if necessary
            if master_patch.dtype != test_patch.dtype:
                test_patch = test_patch.astype(master_patch.dtype)

            # Perform template matching
            result = cv2.matchTemplate(test_patch, master_patch, cv2.TM_CCOEFF_NORMED)
            value = result[0][0]

            def ceil_jutsu(value, decimal_places):
                factor = 10 ** decimal_places
                return math.ceil(value * factor) / factor
            
            match_value = ceil_jutsu(value, 2)

            threshold = 0.96    

            # Determine if the patches match
            if match_value >= threshold:
                print(match_value)
                gray_patch_matches_1.append(True)
            else:
                print(match_value)
                gray_patch_matches_1.append(False)

            # Delete patches to free up memory
            del master_patch, test_patch, result, match_value

            # Optionally, force garbage collection (though usually not necessary)
            gc.collect()

        # Process patches from gray_patch_locations_2
        for i, (x, y, width, height) in enumerate(gray_patch_locations_2):
            # Extract the patch from the original image
            master_patch = extract_patch(reference_image, x, y, width, height)

            # Extract the patch from the new input image
            test_patch = extract_patch(matched, x, y, width, height)

            # Ensure the patches are of the same type and convert if necessary
            if master_patch.dtype != test_patch.dtype:
                test_patch = test_patch.astype(master_patch.dtype)

            # Perform template matching
            result = cv2.matchTemplate(test_patch, master_patch, cv2.TM_CCOEFF_NORMED)
            value = result[0][0]

            match_value = ceil_jutsu(value, 2)

            threshold = 0.69    

            # Determine if the patches match
            if match_value >= threshold:
                print(match_value)
                gray_patch_matches_2.append(True)
            else:
                print(match_value)
                gray_patch_matches_2.append(False)

            # Delete patches to free up memory
            del master_patch, test_patch, result, match_value

            # Optionally, force garbage collection (though usually not necessary)
            gc.collect()

        # Combine patch matching results from both sets
        all_patch_matches = gray_patch_matches_1 + gray_patch_matches_2

        # Ensure the matched image is in a compatible format (uint8)
        matched_uint8 = cv2.normalize(matched, None, 0, 255, cv2.NORM_MINMAX).astype('uint8')

        # Convert the matched image to BGR (color) before drawing rectangles
        result_image = cv2.cvtColor(matched_uint8, cv2.COLOR_GRAY2BGR)

        # Draw rectangles around unmatched patches for gray_patch_locations_1
        for i, (x, y, width, height) in enumerate(gray_patch_locations_1):
            if gray_patch_matches_1[i]:
                color = (0, 255, 0)  # Green for matched patches
            else:
                color = (0, 0, 255)  # Red for unmatched patches
            cv2.rectangle(result_image, (x, y), (x + width, y + height), color, 2)

        # Draw rectangles around unmatched patches for gray_patch_locations_2
        for i, (x, y, width, height) in enumerate(gray_patch_locations_2):
            if gray_patch_matches_2[i]:
                color = (0, 255, 0)  # Green for matched patches
            else:
                color = (0, 0, 255)  # Red for unmatched patches
            cv2.rectangle(result_image, (x, y), (x + width, y + height), color, 2)

        # Display the overall matching status
        status = "OK" if all(all_patch_matches) else "NG"

        # Handle Arduino relay control based on status
        if arduino.is_open:
            try:
                if status == "NG":
                    arduino.write(b'1')  # Turn Relay 1 ON
                    arduino.write(b'3')  # Turn Relay 2 ON
                    time.sleep(5)
                    arduino.write(b'0')  # Turn Relay 1 OFF
                    arduino.write(b'2')  # Turn Relay 2 OFF
                elif status == "OK":
                    arduino.write(b'3')  # Turn Relay 2 ON
                    time.sleep(5)
                    arduino.write(b'2')  # Turn Relay 2 OFF
            except serial.SerialException as e:
                logger.error(f"Failed to communicate with Arduino: {e}")
            finally:
                arduino.close()  # Close the serial connection when done
        else:
            logger.error("Serial port is not open.")

        # Encode the processed image as base64
        _, buffer = cv2.imencode('.jpg', result_image)
        img_base64 = base64.b64encode(buffer).decode('utf-8')

        # Generate inspection report and save the image to the appropriate folder
        inspection_report_path = generate_inspection_report(language, status, result_image, username)
        
        # Return the status of processing along with the processed image
        return jsonify({'status': status, 'image': img_base64, 'inspection_report': inspection_report_path})

    except Exception as e:
        logger.error("Error:", exc_info=True)
        return jsonify({'status': 'Error', 'message': str(e)})


@app.route('/add_rectangle', methods=['POST'])
def add_rectangle():
    try:
        request_data = request.get_json()
        folder_name = request_data['folderName']
        sheet_name = request_data['sheetName']
        rect_data = request_data['data']
        
        folder_path = os.path.join('static', 'language_models', folder_name)
        os.makedirs(folder_path, exist_ok=True)

        file_path = os.path.join(folder_path, 'location_data.xlsx')

        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove the default sheet
            workbook.create_sheet('Sheet1')
            workbook.create_sheet('Sheet2')

        sheet = workbook[sheet_name]

        if sheet.max_row == 1:
            # If the sheet is empty, write the header
            sheet.append(["x", "y", "width", "height"])

        # Write rectangle data
        sheet.append([rect_data['x'], rect_data['y'], rect_data['width'], rect_data['height']])

        workbook.save(file_path)

        return jsonify(success=True)
    except Exception as e:
        logger.error(f"Error adding rectangle: {e}", exc_info=True)
        return jsonify(success=False, message=str(e))

@app.route('/save_all', methods=['POST'])
def save_all():
    try:
        request_data = request.get_json()
        folder_name = request_data['folderName']
        master_name = request_data['masterName']
        date_time = request_data['dateTimeString']

        # Parse the date-time string as UTC
        parsed_date_time = datetime.fromisoformat(date_time[:-1]).replace(tzinfo=timezone.utc)
        
        # Convert UTC to local time by specifying the offset manually (example: +05:30 for IST)
        local_offset = timedelta(hours=5, minutes=30)  # Replace with your local time offset
        local_date_time = parsed_date_time + local_offset

        # Format the date and time
        formatted_date = local_date_time.strftime("%d-%m-%Y")
        formatted_time = local_date_time.strftime("%I:%M %p")
        
        folder_path = os.path.join('static', 'language_models', folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Save the captured image as reference_image.bmp
        captured_image_path = os.path.join(app.config['UPLOAD_FOLDER'], 'captured_image.bmp')
        if os.path.exists(captured_image_path):
            image = Image.open(captured_image_path)
            image_path = os.path.join(folder_path, 'reference_image.bmp')
            image.save(image_path)
        else:
            return jsonify(success=False, message="Captured image not found.")

        # Save the master creator's name and date-time in a text file
        info_file_path = os.path.join(folder_path, 'info.txt')
        write_mode = 'a' if os.path.exists(info_file_path) else 'w'
        with open(info_file_path, write_mode) as f:
            f.write(f"\nMaster Creator: {master_name}\n")
            f.write(f"Date: {formatted_date}\n")
            f.write(f"Time: {formatted_time}\n")
        
        return jsonify(success=True)
    except Exception as e:
        logger.error(f"Error saving all data: {e}", exc_info=True)
        return jsonify(success=False, message=str(e))

# Directory for storing inspection tables
INSPECTION_TABLE_DIR = os.path.join(app.static_folder, 'inspection_table')

if not os.path.exists(INSPECTION_TABLE_DIR):
    os.makedirs(INSPECTION_TABLE_DIR)

# Function to get the current inspection file path
def get_inspection_file_path():
    """Generate the file path for today's inspection data."""
    today = datetime.now().strftime('%d-%m-%Y')
    filename = f"inspection_table_{today}.xlsx"
    return os.path.join('static', 'inspection_table', filename)

# Function to load inspection data from the XLSX file
def load_inspection_data():
    file_path = get_inspection_file_path()
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        return df.to_dict(orient='records')
    else:
        return []

# Function to save inspection data to the XLSX file
def save_inspection_data(inspection_data):
    file_path = get_inspection_file_path()
    df = pd.DataFrame(inspection_data)
    df.to_excel(file_path, index=False)

# Route to fetch inspection data
@app.route('/get_inspection_data')
def get_inspection_data():
    data = load_inspection_data()
    print("Inspection Data Sent to Frontend:", data)  # Debugging line
    return jsonify(data)

# Route to add new inspection record
@app.route('/add_inspection_record', methods=['POST'])
def add_inspection_record():
    data = load_inspection_data()
    new_record = request.json
    data.append(new_record)
    save_inspection_data(data)
    return jsonify({'status': 'success'})

                    
# Run the Flask app
if __name__ == '__main__':
    app.run(debug=True, port=5000)