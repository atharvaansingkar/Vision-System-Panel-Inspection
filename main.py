# Import necessary libraries
from flask import Flask, render_template, request, jsonify, redirect, url_for
import cv2
import numpy as np
import requests
from skimage.exposure import match_histograms
import os
from sklearn.pipeline import Pipeline
import datetime
import base64
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl import Workbook
from PIL import Image
import logging
import neoapi
import pickle
import webview
import threading

#Initialize Flask app
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'static')

# Load the model from file
with open('static/quality_control_logistic_regression_model.pkl', 'rb') as f:
    model = pickle.load(f)

# Define the credentials
USERNAME = 'admin'
PASSWORD = 'admin'

# Initialize logger
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Function to generate JSON inspection report and handle folder creation
def generate_inspection_report(language, status, image):
    try:
        # Create inspection report data
        timestamp_ms = int(datetime.datetime.now().timestamp() * 1000)
        current_date = datetime.datetime.now().strftime("%d-%m-%Y")
        
        # Define the base directory for saving the images
        base_dir = os.path.join(r'C:\Program Files\Apache Software Foundation\Tomcat 9.0\webapps\sfactorystd\Panel_Inspection', current_date)
        
        # Create date folder if it doesn't exist
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        
        # Create model name directory inside date folder
        model_dir = os.path.join(base_dir, language)
        if not os.path.exists(model_dir):
            os.makedirs(model_dir)
        
        # Create OK and NG folders inside model name directory
        ok_dir = os.path.join(model_dir, 'OK')
        ng_dir = os.path.join(model_dir, 'NG')
        if not os.path.exists(ok_dir):
            os.makedirs(ok_dir)
        if not os.path.exists(ng_dir):
            os.makedirs(ng_dir)

        # Save the processed image to the appropriate folder based on status
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        result_filename = f"{language}_{timestamp}.jpg"
        status_folder = 'OK' if status == 'OK' else 'NG'
        image_path = os.path.join(model_dir, status_folder, result_filename)
        cv2.imwrite(image_path, image)

        # Also save the result image as result_image.bmp in the static folder
        static_folder_path = r'D:/pisx_test/static'
        static_image_path = os.path.join(static_folder_path, 'result_image.bmp')
        cv2.imwrite(static_image_path, image)

        # Extract the relative path to be sent in the API request
        relative_path = os.path.relpath(image_path, r'C:\Program Files\Apache Software Foundation\Tomcat 9.0\webapps\sfactorystd')

        # Create the report data
        report_data = {
            "FunctionName": "1007", 
            "Name": "aipl",
            "Password": "aipl",
            "UniqueID": 5,
            "DateTime": str(timestamp_ms),
            "UniqueKey": language,  # Use language as the UniqueKey
            "Command": 3,
            "CommandData": {
                "CommandHeader": [
                    {"headerkey": "ImagePath", "headervalue": relative_path.replace("\\", "/")},
                    {"headerkey": "FinalResult", "headervalue": status}
                ]
            }
        }

        # Send API data to the desired endpoint
        response = requests.post('http://localhost:8080/sfactorystd/', json=report_data)
        if response.status_code == 200:
            return response.text  # Return response if needed
        else:
            return None

    except Exception as e:
        logger.error("Error generating inspection report:", exc_info=True)
        return None

# Load patch locations function
def load_patch_locations(language):
    try:
        base_path = os.path.join('static', 'language_models', language)
        excel_file = os.path.join(base_path, 'location_data.xlsx')
        
        if not os.path.exists(excel_file):
            logger.error(f"Location data file for language {language} does not exist.")
            return []

        wb = openpyxl.load_workbook(excel_file, data_only=True)
        gray_sheet = wb[wb.sheetnames[0]]
        gray_patch_locations = [row[0:4] for row in gray_sheet.iter_rows(min_row=2, values_only=True)]
        
        return gray_patch_locations
    
    except Exception as e:
        logger.error(f"Error loading patch locations for language {language}: {e}", exc_info=True)
        return []
    
# Define the route for capturing image from the camera
@app.route('/capture', methods=['POST'])
def capture_image():
    try:
        camera = neoapi.Cam()
        camera.Connect()
        
        if camera.IsConnected():
            image = camera.GetImage()
            filename = 'captured_image.bmp'
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            image.Save(image_path)
            
            return jsonify({'success': True, 'image_path': f'/static/{filename}'})
        else:
            return jsonify({'success': False, 'error': 'Camera not connected'})
    except Exception as e:
        logger.error(f"Error capturing image: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Failed to capture image'})


# Define the route for the homepage
@app.route('/')
def index():
    return render_template('index.html')

# Define the route for the login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    error_message = None  # Initialize error message
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if username == USERNAME and password == PASSWORD:
            # Redirect to another page upon successful login
            return redirect(url_for('another_page'))
        else:
            error_message = 'Incorrect username or password. Please try again.'

    return render_template('login.html', error_message=error_message)

# Define the route for master page (you can replace this with the desired page)
@app.route('/master')
def another_page():
    return render_template('master.html')

# Define the route for processing images
@app.route('/process', methods=['POST'])
def process_image():
    try:
        # Get selected language from the request
        language = request.form['language']
        
        # Load the uploaded image
        uploaded_image = cv2.imread("static/captured_image.bmp", cv2.IMREAD_COLOR)

        # Construct the dynamic path for the reference image
        reference_image_path = os.path.join('static', 'language_models', language, 'reference_image.bmp')

        # Check if the reference image exists
        if not os.path.isfile(reference_image_path):
            return jsonify({"status": "error", "message": "Reference image not found."}), 404
        
        # Load the reference image
        reference_image = cv2.imread(reference_image_path, cv2.IMREAD_COLOR)

        # Perform histogram matching
        matched = match_histograms(uploaded_image, reference_image, channel_axis=-1)

        # Read the image to be registered
        image_to_register = matched

        # Perform image registration
        sift = cv2.SIFT_create()
        keypoints_ref, descriptors_ref = sift.detectAndCompute(reference_image, None)
        keypoints_to_register, descriptors_to_register = sift.detectAndCompute(image_to_register, None)

        flann = cv2.FlannBasedMatcher()
        matches = flann.knnMatch(descriptors_ref, descriptors_to_register, k=2)

        # Apply ratio test to filter good matches
        good_matches = []
        for m, n in matches:
            if m.distance < 0.75 * n.distance:
                good_matches.append(m)

        # Extract matched keypoints
        matched_keypoints_ref = np.float32([keypoints_ref[m.queryIdx].pt for m in good_matches]).reshape(-1, 1, 2)
        matched_keypoints_to_register = np.float32([keypoints_to_register[m.trainIdx].pt for m in good_matches]).reshape(-1, 1, 2)

        # Find homography
        H, _ = cv2.findHomography(matched_keypoints_to_register, matched_keypoints_ref, cv2.RANSAC, 5.0)

        # Warp the image to be registered to align with the reference image
        registered_image = cv2.warpPerspective(image_to_register, H, (reference_image.shape[1], reference_image.shape[0]))

        # Template Matching
        def extract_patch(image, x, y, width, height):
            """Function to extract a patch from an image based on the given location and size."""
            patch = image[y:y+height, x:x+width]
            return patch

        # Load patch locations based on the selected language
        gray_patch_locations = load_patch_locations(language)

        # Initialize lists to store patch matching results
        gray_patch_matches = []

        # Read the original image using OpenCV for grayscale image processing
        sample = reference_image

        # Convert the original image to grayscale using OpenCV
        sample_g = cv2.cvtColor(sample, cv2.COLOR_BGR2GRAY)

        # Extract the scaler from the model pipeline
        if isinstance(model, Pipeline):
            scaler = model.named_steps['scaler']
            clf = model.named_steps['logreg']
        else:
            raise ValueError("Model is not a pipeline. This approach assumes a pipeline with scaler and logistic regression.")

        # Iterate over each patch location and size for grayscale image processing
        for i, (x, y, width, height) in enumerate(gray_patch_locations):
            # Extract the patch from the original image
            patch_original = extract_patch(sample_g, x, y, width, height)

            # Read the new input image using OpenCV
            new_sample_g = registered_image

            # Convert the registered image to grayscale
            new_sample_g = cv2.cvtColor(registered_image, cv2.COLOR_BGR2GRAY)

            # Extract the patch from the new input image
            new_patch = extract_patch(new_sample_g, x, y, width, height)

            # Perform template matching between the patches
            Value = cv2.matchTemplate(new_patch, patch_original, cv2.TM_CCOEFF_NORMED)
            _, max_val, _, _ = cv2.minMaxLoc(Value)

            # Prepare input for the model
            prediction_input = np.array([[max_val]])  # This should be a 2D array

            # Scale the input
            scaled_input = scaler.transform(prediction_input)  # Remove check_input argument

            # Predict using the model
            prediction = clf.predict(scaled_input)[0]

            # Determine if the patches match
            if scaled_input > 0.725:
                gray_patch_matches.append(True)
                print(scaled_input)
            else:
                print(scaled_input)
                gray_patch_matches.append(False)

        # Combine patch matching results from both color and grayscale processing
        all_patch_matches = gray_patch_matches

        # Draw rectangles around unmatched patches
        for i, (x, y, width, height) in enumerate(gray_patch_locations):
            if all_patch_matches[i]:
                color = (0, 255, 0)  # Green for matched patches
            else:
                color = (0, 0, 255)  # Red for unmatched patches
            cv2.rectangle(registered_image, (x, y), (x + width, y + height), color, 2)

        # Display the overall matching status
        status = "OK" if all(all_patch_matches) else "NG"

        # Encode the processed image as base64
        _, buffer = cv2.imencode('.jpg', registered_image)
        img_base64 = base64.b64encode(buffer).decode('utf-8')

        # Generate inspection report and save the image to the appropriate folder
        inspection_report_path = generate_inspection_report(language, status, registered_image)
        
        # Return the status of processing along with the processed image
        return jsonify({'status': status, 'image': img_base64, 'inspection_report': inspection_report_path})

    except Exception as e:
        logger.error("Error:", exc_info=True)
        return jsonify({'status': 'Error', 'message': str(e)})

    
@app.route('/save_rectangles', methods=['POST'])
def save_rectangles():
    try:
        request_data = request.get_json()
        folder_name = request_data['folderName']
        data = request_data['data']
        
        folder_path = os.path.join('static', 'language_models', folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Save the captured image as reference_image.bmp
        captured_image_path = os.path.join(app.config['UPLOAD_FOLDER'], 'captured_image.bmp')
        if os.path.exists(captured_image_path):
            image = Image.open(captured_image_path)
            image_path = os.path.join(folder_path, 'reference_image.bmp')
            image.save(image_path)
        else:
            return jsonify(success=False, message="Captured image not found.")

        # Save the rectangles data in an Excel file
        file_path = os.path.join(folder_path, 'location_data.xlsx')
        workbook = Workbook()
        sheet = workbook.active

        # Write the header
        sheet.append(["x", "y", "width", "height"])

        # Write rectangle data
        for rect in data:
            sheet.append([rect['x'], rect['y'], rect['width'], rect['height']])

        workbook.save(file_path)
        return jsonify(success=True)
    except Exception as e:
        logger.error(f"Error saving rectangles: {e}", exc_info=True)
        return jsonify(success=False, message=str(e))  

def start_flask():
    app.run()

if __name__ == '__main__':
    flask_thread = threading.Thread(target=start_flask)
    flask_thread.daemon = True
    flask_thread.start()
    
    webview.create_window("Vision System Panel Inspection", "http://127.0.0.1:5000/")
    webview.start()
